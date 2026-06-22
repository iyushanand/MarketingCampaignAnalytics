import sys
import os
import site

# Force inclusion of user site-packages in case it's disabled by the hosting server
user_site = site.getusersitepackages()
if os.path.exists(user_site) and user_site not in sys.path:
    sys.path.insert(0, user_site)

import json
import datetime
import warnings
warnings.filterwarnings("ignore")

# openpyxl imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# reportlab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Numbered Canvas for PDF page headers and footers
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_decorations(self, page_count):
        self.saveState()
        # Draw header
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#1F4E79"))
        self.drawString(54, 755, "MARKETING CAMPAIGN ANALYTICS PLATFORM")
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#595959"))
        self.drawRightString(558, 755, "EXECUTIVE PERFORMANCE REPORT")
        
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.75)
        self.line(54, 750, 558, 750)
        
        # Draw footer
        self.line(54, 50, 558, 50)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 35, page_text)
        gen_date = datetime.datetime.now().strftime("%B %d, %Y")
        self.drawString(54, 35, f"Confidential  |  Generated on {gen_date}")
        self.restoreState()

def create_excel_report(output_path, data):
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Executive Summary
    # ----------------------------------------------------
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_exec.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    kpi_val_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title
    ws_exec["A1"] = "Executive Marketing Performance Summary"
    ws_exec["A1"].font = title_font
    ws_exec.row_dimensions[1].height = 30
    
    # KPI cards layout
    kpis = [
        ("Total Revenue", data["executiveSummary"]["totalRevenue"], "$#,##0.00"),
        ("Total Spend", data["executiveSummary"]["totalSpend"], "$#,##0.00"),
        ("Return on Investment (ROI)", data["executiveSummary"]["roi"] / 100.0, "0.00%"),
        ("Avg Response Rate", data["executiveSummary"]["averageResponseRate"] / 100.0, "0.00%"),
        ("Avg Conversion Rate", data["executiveSummary"]["averageConversionRate"] / 100.0, "0.00%"),
    ]
    
    for idx, (label, val, fmt) in enumerate(kpis):
        col_lbl = get_column_letter(idx + 1)
        # Label cell (Row 3)
        lbl_cell = ws_exec[f"{col_lbl}3"]
        lbl_cell.value = label
        lbl_cell.font = label_font
        lbl_cell.fill = accent_fill
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lbl_cell.border = thin_border
        
        # Value cell (Row 4)
        val_cell = ws_exec[f"{col_lbl}4"]
        val_cell.value = val
        val_cell.font = kpi_val_font
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.number_format = fmt
        val_cell.border = thin_border
        
    ws_exec.row_dimensions[3].height = 25
    ws_exec.row_dimensions[4].height = 30
    
    # Additional Quick Summary text
    ws_exec["A6"] = "Best Performing Campaign:"
    ws_exec["A6"].font = label_font
    ws_exec["B6"] = data["executiveSummary"]["bestCampaign"]
    ws_exec["B6"].font = normal_font
    
    ws_exec["A7"] = "Highest Performing Channel:"
    ws_exec["A7"].font = label_font
    ws_exec["B7"] = data["executiveSummary"]["bestChannel"]
    ws_exec["B7"].font = normal_font
    
    # Write channels revenue mini table for Pie Chart reference (placed at G4:H9)
    ws_exec["G3"] = "Marketing Channel"
    ws_exec["G3"].font = header_font
    ws_exec["G3"].fill = header_fill
    ws_exec["G3"].border = thin_border
    
    ws_exec["H3"] = "Total Revenue"
    ws_exec["H3"].font = header_font
    ws_exec["H3"].fill = header_fill
    ws_exec["H3"].border = thin_border
    
    # Extract channel revenues from campaign performance
    channel_rev = {}
    for camp in data["campaignPerformance"]:
        ch = camp["channel"]
        channel_rev[ch] = channel_rev.get(ch, 0.0) + float(camp["revenue"])
        
    row_idx = 4
    for ch, rev in channel_rev.items():
        ws_exec[f"G{row_idx}"] = ch
        ws_exec[f"G{row_idx}"].font = normal_font
        ws_exec[f"G{row_idx}"].border = thin_border
        
        ws_exec[f"H{row_idx}"] = rev
        ws_exec[f"H{row_idx}"].font = normal_font
        ws_exec[f"H{row_idx}"].number_format = "$#,##0.00"
        ws_exec[f"H{row_idx}"].border = thin_border
        row_idx += 1
        
    # Add Pie Chart for Channel Revenue
    pie = PieChart()
    pie.title = "Revenue Contribution by Marketing Channel"
    pie.style = 10
    
    data_ref = Reference(ws_exec, min_col=8, min_row=3, max_row=row_idx - 1)
    cats_ref = Reference(ws_exec, min_col=7, min_row=4, max_row=row_idx - 1)
    
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws_exec.add_chart(pie, "A9")
    
    # ----------------------------------------------------
    # Sheet 2: Campaign Performance
    # ----------------------------------------------------
    ws_camp = wb.create_sheet(title="Campaign Performance")
    ws_camp.views.sheetView[0].showGridLines = True
    
    ws_camp["A1"] = "Detailed Campaign Performance Analysis"
    ws_camp["A1"].font = title_font
    ws_camp.row_dimensions[1].height = 25
    
    # Table headers
    headers = [
        "Campaign Name", "Marketing Channel", "Campaign Spend", 
        "Revenue Generated", "Return on Investment (ROI)", "ROAS", 
        "Conversions", "Response Rate"
    ]
    
    for idx, h in enumerate(headers):
        cell = ws_camp.cell(row=3, column=idx + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_camp.row_dimensions[3].height = 28
    
    # Write rows
    start_row = 4
    for idx, c in enumerate(data["campaignPerformance"]):
        curr_row = start_row + idx
        ws_camp.cell(row=curr_row, column=1, value=c["campaignName"]).font = normal_font
        ws_camp.cell(row=curr_row, column=2, value=c["channel"]).font = normal_font
        
        c_spend = ws_camp.cell(row=curr_row, column=3, value=float(c["spend"]))
        c_spend.font = normal_font
        c_spend.number_format = "$#,##0.00"
        
        c_rev = ws_camp.cell(row=curr_row, column=4, value=float(c["revenue"]))
        c_rev.font = normal_font
        c_rev.number_format = "$#,##0.00"
        
        c_roi = ws_camp.cell(row=curr_row, column=5, value=float(c["roi"]) / 100.0)
        c_roi.font = normal_font
        c_roi.number_format = "0.00%"
        
        c_roas = ws_camp.cell(row=curr_row, column=6, value=float(c["roas"]))
        c_roas.font = normal_font
        c_roas.number_format = "0.00"
        
        c_conv = ws_camp.cell(row=curr_row, column=7, value=int(c["conversions"]))
        c_conv.font = normal_font
        c_conv.number_format = "#,##0"
        
        c_resp = ws_camp.cell(row=curr_row, column=8, value=float(c["responseRate"]) / 100.0)
        c_resp.font = normal_font
        c_resp.number_format = "0.00%"
        
        # Borders and zebra stripping
        for col_idx in range(1, 9):
            cell = ws_camp.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = accent_fill
                
        ws_camp.row_dimensions[curr_row].height = 20
        
    last_camp_row = start_row + len(data["campaignPerformance"]) - 1
    
    # Conditional formatting rules
    # 1. Negative ROI -> Red
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    ws_camp.conditional_formatting.add(
        f"E4:E{last_camp_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
    )
    
    # 2. Highest Revenue -> Green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    max_rev_rule = FormulaRule(formula=[f"D4=MAX($D$4:$D${last_camp_row})"], fill=green_fill, font=green_font)
    ws_camp.conditional_formatting.add(f"D4:D{last_camp_row}", max_rev_rule)
    
    # 3. Highest Response Rate -> Blue
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    blue_font = Font(color="1F4E79")
    max_resp_rule = FormulaRule(formula=[f"H4=MAX($H$4:$H${last_camp_row})"], fill=blue_fill, font=blue_font)
    ws_camp.conditional_formatting.add(f"H4:H{last_camp_row}", max_resp_rule)
    
    # Freeze row headers
    ws_camp.freeze_panes = "A4"
    
    # Add Bar Chart
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Revenue by Campaign ($)"
    bar.y_axis.title = "Revenue"
    bar.x_axis.title = "Campaign"
    
    data_ref = Reference(ws_camp, min_col=4, min_row=3, max_row=last_camp_row)
    cats_ref = Reference(ws_camp, min_col=1, min_row=4, max_row=last_camp_row)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.legend = None
    
    ws_camp.add_chart(bar, "A12")
    
    # ----------------------------------------------------
    # Sheet 3: Customer Analytics
    # ----------------------------------------------------
    ws_cust = wb.create_sheet(title="Customer Analytics")
    ws_cust.views.sheetView[0].showGridLines = True
    
    ws_cust["A1"] = "Customer Segmentation Summary"
    ws_cust["A1"].font = title_font
    ws_cust.row_dimensions[1].height = 25
    
    cust_headers = ["Customer Segment", "Customer Count", "Average Spend", "Average Income", "Average Purchases", "Response Rate"]
    for idx, h in enumerate(cust_headers):
        cell = ws_cust.cell(row=3, column=idx + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_cust.row_dimensions[3].height = 28
    
    for idx, s in enumerate(data["customerAnalytics"]):
        curr_row = start_row + idx
        ws_cust.cell(row=curr_row, column=1, value=s["segment"]).font = normal_font
        
        c_count = ws_cust.cell(row=curr_row, column=2, value=int(s["customerCount"]))
        c_count.font = normal_font
        c_count.number_format = "#,##0"
        
        c_spend = ws_cust.cell(row=curr_row, column=3, value=float(s["averageSpend"]))
        c_spend.font = normal_font
        c_spend.number_format = "$#,##0.00"
        
        c_inc = ws_cust.cell(row=curr_row, column=4, value=float(s["averageIncome"]))
        c_inc.font = normal_font
        c_inc.number_format = "$#,##0.00"
        
        c_pur = ws_cust.cell(row=curr_row, column=5, value=float(s["averagePurchases"]))
        c_pur.font = normal_font
        c_pur.number_format = "0.0"
        
        c_resp = ws_cust.cell(row=curr_row, column=6, value=float(s["responseRate"]) / 100.0)
        c_resp.font = normal_font
        c_resp.number_format = "0.00%"
        
        for col_idx in range(1, 7):
            cell = ws_cust.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = accent_fill
                
        ws_cust.row_dimensions[curr_row].height = 20
        
    ws_cust.freeze_panes = "A4"
    
    # ----------------------------------------------------
    # Sheet 4: Monthly Revenue
    # ----------------------------------------------------
    ws_month = wb.create_sheet(title="Monthly Revenue")
    ws_month.views.sheetView[0].showGridLines = True
    
    ws_month["A1"] = "Monthly Sales & Investment Trend"
    ws_month["A1"].font = title_font
    ws_month.row_dimensions[1].height = 25
    
    month_headers = ["Month", "Revenue Generated", "Campaign Spend", "Monthly ROI"]
    for idx, h in enumerate(month_headers):
        cell = ws_month.cell(row=3, column=idx + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_month.row_dimensions[3].height = 28
    
    for idx, m in enumerate(data["monthlyRevenue"]):
        curr_row = start_row + idx
        ws_month.cell(row=curr_row, column=1, value=m["month"]).font = normal_font
        
        c_rev = ws_month.cell(row=curr_row, column=2, value=float(m["revenue"]))
        c_rev.font = normal_font
        c_rev.number_format = "$#,##0.00"
        
        c_spend = ws_month.cell(row=curr_row, column=3, value=float(m["spend"]))
        c_spend.font = normal_font
        c_spend.number_format = "$#,##0.00"
        
        c_roi = ws_month.cell(row=curr_row, column=4, value=float(m["roi"]) / 100.0)
        c_roi.font = normal_font
        c_roi.number_format = "0.00%"
        
        for col_idx in range(1, 5):
            cell = ws_month.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = accent_fill
                
        ws_month.row_dimensions[curr_row].height = 20
        
    last_month_row = start_row + len(data["monthlyRevenue"]) - 1
    ws_month.freeze_panes = "A4"
    
    # Add Line Chart
    line = LineChart()
    line.title = "Monthly Revenue Trend ($)"
    line.style = 13
    line.y_axis.title = "Revenue"
    line.x_axis.title = "Month"
    
    data_ref = Reference(ws_month, min_col=2, min_row=3, max_row=last_month_row)
    cats_ref = Reference(ws_month, min_col=1, min_row=4, max_row=last_month_row)
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    line.legend = None
    
    ws_month.add_chart(line, "F3")
    
    # ----------------------------------------------------
    # Auto-adjust column widths across all sheets
    # ----------------------------------------------------
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('$' in cell.number_format or '%' in cell.number_format):
                    max_len = max(max_len, len(val) + 6)
                else:
                    max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    wb.save(output_path)

def create_pdf_report(output_path, data):
    # Base setup
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette
    color_primary = colors.HexColor("#1F4E79")
    color_secondary = colors.HexColor("#595959")
    color_light_bg = colors.HexColor("#F2F7FA")
    
    # Modify default styles and create custom ones
    style_title = ParagraphStyle(
        'CoverTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=color_primary,
        spaceAfter=10
    )
    
    style_subtitle = ParagraphStyle(
        'CoverSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=color_secondary,
        spaceAfter=30
    )
    
    style_h1 = ParagraphStyle(
        'HeaderH1',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=color_primary,
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    style_body = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#333333"),
        spaceAfter=10
    )
    
    style_table_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    style_table_cell = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#333333")
    )
    
    style_table_cell_bold = ParagraphStyle(
        'TableCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#333333")
    )
    
    story = []
    
    # ----------------------------------------------------
    # Header / Title Block
    # ----------------------------------------------------
    story.append(Paragraph("Executive Marketing Analysis Report", style_title))
    story.append(Paragraph(f"Corporate Performance Review  |  Data Active Baseline: 2026", style_subtitle))
    story.append(Spacer(1, 10))
    
    # ----------------------------------------------------
    # Section 1: Executive Summary
    # ----------------------------------------------------
    story.append(Paragraph("1. Executive Summary", style_h1))
    
    exec_summary_text = (
        f"This business intelligence report delivers an automated performance evaluation of campaign investments "
        f"and customer demographic distributions. The complete marketing portfolio generated a total revenue of "
        f"<b>${data['executiveSummary']['totalRevenue']:,.2f}</b> on an aggregated campaign spend of "
        f"<b>${data['executiveSummary']['totalSpend']:,.2f}</b>, resulting in a net return on investment (ROI) of "
        f"<b>{data['executiveSummary']['roi']:.2f}%</b>. The average customer acquisition funnel registered a response "
        f"rate of <b>{data['executiveSummary']['averageResponseRate']:.2f}%</b> and a conversion rate of "
        f"<b>{data['executiveSummary']['averageConversionRate']:.2f}%</b>. The highest performing campaign was "
        f"<b>{data['executiveSummary']['bestCampaign']}</b>, with the most lucrative outreach channel identified as "
        f"<b>{data['executiveSummary']['bestChannel']}</b>."
    )
    story.append(Paragraph(exec_summary_text, style_body))
    story.append(Spacer(1, 10))
    
    # KPI Grid Table
    kpi_data = [
        [
            Paragraph("<b>Total Revenue</b>", style_table_cell_bold),
            Paragraph("<b>Total Campaign Spend</b>", style_table_cell_bold),
            Paragraph("<b>Portfolio ROI</b>", style_table_cell_bold),
            Paragraph("<b>Avg Response Rate</b>", style_table_cell_bold)
        ],
        [
            Paragraph(f"${data['executiveSummary']['totalRevenue']:,.2f}", style_table_cell),
            Paragraph(f"${data['executiveSummary']['totalSpend']:,.2f}", style_table_cell),
            Paragraph(f"{data['executiveSummary']['roi']:.2f}%", style_table_cell),
            Paragraph(f"{data['executiveSummary']['averageResponseRate']:.2f}%", style_table_cell)
        ]
    ]
    
    t_kpi = Table(kpi_data, colWidths=[126, 126, 126, 126])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), color_light_bg),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 15))
    
    # ----------------------------------------------------
    # Section 2: Campaign KPI Table
    # ----------------------------------------------------
    story.append(Paragraph("2. Campaign Performance Overview", style_h1))
    
    camp_table_data = [[
        Paragraph("Campaign Name", style_table_header),
        Paragraph("Channel", style_table_header),
        Paragraph("Spend", style_table_header),
        Paragraph("Revenue", style_table_header),
        Paragraph("ROI", style_table_header),
        Paragraph("ROAS", style_table_header),
        Paragraph("Conversions", style_table_header),
        Paragraph("Resp. Rate", style_table_header)
    ]]
    
    for c in data["campaignPerformance"]:
        camp_table_data.append([
            Paragraph(c["campaignName"], style_table_cell_bold),
            Paragraph(c["channel"], style_table_cell),
            Paragraph(f"${float(c['spend']):,.2f}", style_table_cell),
            Paragraph(f"${float(c['revenue']):,.2f}", style_table_cell),
            Paragraph(f"{float(c['roi']):.2f}%", style_table_cell),
            Paragraph(f"{float(c['roas']):.2f}", style_table_cell),
            Paragraph(f"{int(c['conversions']):,}", style_table_cell),
            Paragraph(f"{float(c['responseRate']):.2f}%", style_table_cell)
        ])
        
    t_camp = Table(camp_table_data, colWidths=[100, 60, 65, 70, 50, 45, 60, 54])
    t_camp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), color_primary),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ('ALIGN', (2,1), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, color_light_bg])
    ]))
    story.append(t_camp)
    story.append(Spacer(1, 15))
    
    # ----------------------------------------------------
    # Section 3: Customer Summary
    # ----------------------------------------------------
    story.append(Paragraph("3. Customer Segment Analytics", style_h1))
    
    cust_table_data = [[
        Paragraph("Customer Segment", style_table_header),
        Paragraph("Customer Count", style_table_header),
        Paragraph("Average Spend", style_table_header),
        Paragraph("Average Income", style_table_header),
        Paragraph("Average Purchases", style_table_header),
        Paragraph("Response Rate", style_table_header)
    ]]
    
    for s in data["customerAnalytics"]:
        cust_table_data.append([
            Paragraph(s["segment"], style_table_cell_bold),
            Paragraph(f"{int(s['customerCount']):,}", style_table_cell),
            Paragraph(f"${float(s['averageSpend']):,.2f}", style_table_cell),
            Paragraph(f"${float(s['averageIncome']):,.2f}", style_table_cell),
            Paragraph(f"{float(s['averagePurchases']):.1f}", style_table_cell),
            Paragraph(f"{float(s['responseRate']):.2f}%", style_table_cell)
        ])
        
    t_cust = Table(cust_table_data, colWidths=[120, 75, 75, 80, 75, 79])
    t_cust.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), color_primary),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, color_light_bg])
    ]))
    story.append(t_cust)
    story.append(Spacer(1, 15))
    
    # ----------------------------------------------------
    # Section 4: Marketing Insights
    # ----------------------------------------------------
    story.append(Paragraph("4. Marketing Insights", style_h1))
    
    insights_text = (
        "Analysis of demographic cohorts indicates that high-income customer segments are strongly correlated "
        "with loyalty and acquisition responses. In contrast, younger demographic clusters (ages under 30) exhibit "
        "lower conversion volumes. Channels such as Email and Google Search generate the highest return on spend, "
        "whereas Social Media campaigns show a higher cost-per-acquisition. Reallocating marketing budgets toward "
        "the high-value segments and utilizing attribution scoring is recommended."
    )
    story.append(Paragraph(insights_text, style_body))
    story.append(Spacer(1, 10))
    
    # ----------------------------------------------------
    # Section 5: Business Recommendations
    # ----------------------------------------------------
    story.append(Paragraph("5. Business Recommendations", style_h1))
    
    recs = []
    best_ch = data["executiveSummary"]["bestChannel"]
    recs.append(f"<b>Increase investment in {best_ch} campaigns</b>: This channel historically exhibits the highest conversion rates and ROAS.")
    
    lowest_roi_camp = min(data["campaignPerformance"], key=lambda x: float(x["roi"]))
    if float(lowest_roi_camp["roi"]) < 10.0:
        recs.append(f"<b>Review budget and design for low-performing campaigns</b>: {lowest_roi_camp['campaignName']} generated a low ROI of {float(lowest_roi_camp['roi']):.2f}%. Audit ad placement and audience targeting.")
    
    recs.append("<b>Focus retention and loyalty efforts on High Value Customers</b>: High-value segments represent the highest monetary spend and purchase frequency. Implement high-tier reward programs.")
    recs.append("<b>Optimize campaigns with low response rates</b>: Audit CTR and design funnels to run A/B testing on email subjects and landing page elements to lift user responsiveness.")
    recs.append("<b>Reallocate budget from low-ROI channels</b>: Reassign capital away from channels that consistently produce ROI margins below 10% to secure higher overall return on spend.")
    
    for idx, r in enumerate(recs):
        story.append(Paragraph(f"{idx+1}. {r}", style_body))
        
    doc.build(story, canvasmaker=NumberedCanvas)

def main():
    if len(sys.argv) < 4:
        print("Usage: python report_generator.py <excel|pdf> <output_path> <data_json_path>")
        sys.exit(1)
        
    command = sys.argv[1].lower()
    output_path = sys.argv[2]
    data_json_path = sys.argv[3]
    
    if not os.path.exists(data_json_path):
        print(f"Data JSON file not found: {data_json_path}")
        sys.exit(1)
        
    with open(data_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    target_dir = os.path.dirname(output_path)
    if target_dir and not os.path.exists(target_dir):
        os.makedirs(target_dir, exist_ok=True)
        
    if command == "excel":
        create_excel_report(output_path, data)
        print(f"Excel report generated successfully at: {output_path}")
    elif command == "pdf":
        create_pdf_report(output_path, data)
        print(f"PDF report generated successfully at: {output_path}")
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)

if __name__ == "__main__":
    main()
